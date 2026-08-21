#!/usr/bin/env python3
from __future__ import annotations

# One-time bootstrap only. Remove after the valid editable Excel master is committed.
import base64
import hashlib
import json
import zlib
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
CHUNKS_DIR = ROOT / "scripts/bootstrap_chunks"
OUTPUT = ROOT / "docs/data/chimera_battle_ctb_redesign_v02_clean.xlsx"
EXPECTED_LENGTH = 13524
EXPECTED_SHA256 = "5da6d744486a25b31d78bf7cc0b49776f12ee95921d6a3380378746578a92647"


def load_payload() -> dict:
    chunk_paths = sorted(CHUNKS_DIR.glob("part*.txt"))
    if len(chunk_paths) != 8:
        raise ValueError(f"expected 8 bootstrap chunks, found {len(chunk_paths)}")

    encoded = "".join(path.read_text(encoding="utf-8").strip() for path in chunk_paths)
    if len(encoded) != EXPECTED_LENGTH:
        raise ValueError(f"bootstrap length mismatch: {len(encoded)} != {EXPECTED_LENGTH}")

    digest = hashlib.sha256(encoded.encode("ascii")).hexdigest()
    if digest != EXPECTED_SHA256:
        raise ValueError(f"bootstrap SHA-256 mismatch: {digest} != {EXPECTED_SHA256}")

    raw = zlib.decompress(base64.b64decode(encoded, validate=True))
    payload = json.loads(raw.decode("utf-8"))

    expected = {
        "コマンド": 61,      # header + 60 rows
        "部位": 81,          # header + 80 rows
        "敵": 46,            # header + 45 rows
        "シナジー": 37,      # header + 36 rows
        "状態異常": 25,      # header + 24 rows
        "代表ビルド": 21,    # header + 20 rows
    }
    for sheet, expected_rows in expected.items():
        actual = len(payload["sheets"].get(sheet, []))
        if actual != expected_rows:
            raise ValueError(f"bootstrap row count mismatch: {sheet}: {actual} != {expected_rows}")
    return payload


def build_workbook(payload: dict) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, rows in payload["sheets"].items():
        ws = wb.create_sheet(sheet_name)
        for row in rows:
            ws.append(row)

        ws.freeze_panes = "A2" if len(rows) > 1 else None
        ws.auto_filter.ref = ws.dimensions if len(rows) > 1 else None

        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="404040")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for col_idx in range(1, ws.max_column + 1):
            values = [str(ws.cell(r, col_idx).value or "") for r in range(1, min(ws.max_row, 80) + 1)]
            width = min(max(10, max((len(v) for v in values), default=10) + 2), 34)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT} ({OUTPUT.stat().st_size} bytes)")


if __name__ == "__main__":
    build_workbook(load_payload())
